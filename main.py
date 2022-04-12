# ...
# INSTALL
# pip install openpyxl
# pip install PyQt5
# COMPILE
# pyinstaller -F -w main.py
# ...

import os
import sys
import time
import PyQt5
import PyQt5.QtWidgets
import PyQt5.QtCore
import PyQt5.QtGui
import openpyxl
import openpyxl.utils
import random


# функция преобразования первых трёх ячеек в "ФИО маленького регистра"
def get_fio_low_case(list_in):
    fio_low_case = ''
    for counter in range(0, 3):
        fio_low_case = fio_low_case + ''.join(str(list_in[counter]).lower().split())
    return fio_low_case


# класс главного окна
class Window(PyQt5.QtWidgets.QMainWindow):
    # описание главного окна
    def __init__(self):
        super(Window, self).__init__()

        # переменные, атрибуты
        self.info_for_open_file = ''
        self.info_path_open_file = ''
        self.info_extention_open_file = 'Файлы Excel xlsx (*.xlsx)'
        self.text_empty_path_file = 'файл пока не выбран'

        # TODO
        # заменить эту переменную на 0 или пустоту
        self.max_string = '260'

        # начало диапазона поиска строк в обоих файлах
        self.range_all_files = 'A2:'

        # главное окно, надпись на нём и размеры
        self.setWindowTitle('Добор в эксель')
        self.setGeometry(600, 200, 700, 610)

        # объекты на форме
        # label_full_file
        self.label_full_file = PyQt5.QtWidgets.QLabel(self)
        self.label_full_file.setObjectName('label_full_file')
        self.label_full_file.setText('1. Выберите Полный файл')
        self.label_full_file.setGeometry(PyQt5.QtCore.QRect(10, 10, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_full_file.setFont(font)
        self.label_full_file.adjustSize()
        self.label_full_file.setToolTip(self.label_full_file.objectName())

        # toolButton_select_full_file
        self.toolButton_select_full_file = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_full_file.setObjectName('toolButton_select_full_file')
        self.toolButton_select_full_file.setText('...')
        self.toolButton_select_full_file.setGeometry(PyQt5.QtCore.QRect(10, 40, 50, 20))
        self.toolButton_select_full_file.setFixedWidth(50)
        self.toolButton_select_full_file.clicked.connect(self.select_file)
        self.toolButton_select_full_file.setToolTip(self.toolButton_select_full_file.objectName())

        # label_path_full_file
        self.label_path_full_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_full_file.setObjectName('label_path_full_file')
        self.label_path_full_file.setText(self.text_empty_path_file)
        self.label_path_full_file.setGeometry(PyQt5.QtCore.QRect(70, 40, 820, 16))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_full_file.setFont(font)
        self.label_path_full_file.adjustSize()
        self.label_path_full_file.setToolTip(self.label_path_full_file.objectName())

        # label_half_file
        self.label_half_file = PyQt5.QtWidgets.QLabel(self)
        self.label_half_file.setObjectName('label_half_file')
        self.label_half_file.setText('2. Выберите Неполный файл')
        self.label_half_file.setGeometry(PyQt5.QtCore.QRect(10, 70, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_half_file.setFont(font)
        self.label_half_file.adjustSize()
        self.label_half_file.setToolTip(self.label_half_file.objectName())

        # toolButton_select_half_file
        self.toolButton_select_half_file = PyQt5.QtWidgets.QPushButton(self)
        self.toolButton_select_half_file.setObjectName('toolButton_select_half_file')
        self.toolButton_select_half_file.setText('...')
        self.toolButton_select_half_file.setGeometry(PyQt5.QtCore.QRect(10, 100, 50, 20))
        self.toolButton_select_half_file.setFixedWidth(50)
        self.toolButton_select_half_file.clicked.connect(self.select_file)
        self.toolButton_select_half_file.setToolTip(self.toolButton_select_half_file.objectName())

        # label_path_half_file
        self.label_path_half_file = PyQt5.QtWidgets.QLabel(self)
        self.label_path_half_file.setObjectName('label_path_half_file')
        self.label_path_half_file.setText(self.text_empty_path_file)
        self.label_path_half_file.setGeometry(PyQt5.QtCore.QRect(70, 100, 820, 20))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.label_path_half_file.setFont(font)
        self.label_path_half_file.adjustSize()
        self.label_path_half_file.setToolTip(self.label_path_half_file.objectName())

        # label_max_string
        self.label_max_string = PyQt5.QtWidgets.QLabel(self)
        self.label_max_string.setObjectName('label_max_string')
        self.label_max_string.setText('3. Сколько должно быть строк в файле')
        self.label_max_string.setGeometry(PyQt5.QtCore.QRect(10, 130, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_max_string.setFont(font)
        self.label_max_string.adjustSize()
        self.label_max_string.setToolTip(self.label_max_string.objectName())

        # lineEdit_max_string
        self.lineEdit_max_string = PyQt5.QtWidgets.QLineEdit(self)
        self.lineEdit_max_string.setObjectName('lineEdit_max_string')
        self.lineEdit_max_string.setText(self.max_string)
        self.lineEdit_max_string.setGeometry(PyQt5.QtCore.QRect(10, 160, 90, 20))
        self.lineEdit_max_string.setClearButtonEnabled(True)
        self.lineEdit_max_string.setToolTip(self.lineEdit_max_string.objectName())

        # label_spec_string
        self.label_spec_string = PyQt5.QtWidgets.QLabel(self)
        self.label_spec_string.setObjectName('label_spec_string')
        self.label_spec_string.setText('4. Выберите специальности, одну или несколько')
        self.label_spec_string.setGeometry(PyQt5.QtCore.QRect(10, 190, 150, 40))
        font = PyQt5.QtGui.QFont()
        font.setPointSize(12)
        self.label_spec_string.setFont(font)
        self.label_spec_string.adjustSize()
        self.label_spec_string.setToolTip(self.label_spec_string.objectName())

        # listWidget_specialization
        self.listWidget_specialization = PyQt5.QtWidgets.QListWidget(self)
        self.listWidget_specialization.setObjectName('listWidget_specialization')
        self.listWidget_specialization.setGeometry(PyQt5.QtCore.QRect(10, 220, 400, 300))
        self.listWidget_specialization.setSelectionMode(PyQt5.QtWidgets.QListView.MultiSelection)
        font = PyQt5.QtGui.QFont()
        font.setPointSize(10)
        self.listWidget_specialization.setFont(font)
        self.listWidget_specialization.setResizeMode(PyQt5.QtWidgets.QListView.Adjust)
        self.listWidget_specialization.setEnabled(False)

        # pushButton_do_fill_data
        self.pushButton_do_fill_data = PyQt5.QtWidgets.QPushButton(self)
        self.pushButton_do_fill_data.setObjectName('pushButton_do_fill_data')
        self.pushButton_do_fill_data.setEnabled(False)
        self.pushButton_do_fill_data.setText('Произвести заполнение')
        self.pushButton_do_fill_data.setGeometry(PyQt5.QtCore.QRect(10, 535, 180, 25))
        self.pushButton_do_fill_data.setFixedWidth(130)
        self.pushButton_do_fill_data.clicked.connect(self.do_fill_data)
        self.pushButton_do_fill_data.setToolTip(self.pushButton_do_fill_data.objectName())

        # button_exit
        self.button_exit = PyQt5.QtWidgets.QPushButton(self)
        self.button_exit.setObjectName('button_exit')
        self.button_exit.setText('Выход')
        self.button_exit.setGeometry(PyQt5.QtCore.QRect(10, 570, 180, 25))
        self.button_exit.setFixedWidth(50)
        self.button_exit.clicked.connect(self.click_on_btn_exit)
        self.button_exit.setToolTip(self.button_exit.objectName())

    # событие - нажатие на кнопку выбора файла
    def select_file(self):
        # запоминание старого значения пути выбора файлов
        old_path_of_selected_full_file = self.label_path_full_file.text()
        old_path_of_selected_half_file = self.label_path_half_file.text()

        # определение какая кнопка выбора файла нажата
        if self.sender().objectName() == self.toolButton_select_full_file.objectName():
            self.info_for_open_file = 'Выберите Полный файл формата Excel, версии старше 2007 года (.XLSX)'
        elif self.sender().objectName() == self.toolButton_select_half_file.objectName():
            self.info_for_open_file = 'Выберите Неполный файл формата Excel, версии старше 2007 года (.XLSX)'

        # непосредственное окно выбора файла и переменная для хранения пути файла
        data_of_open_file_name = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self,
                                                                             self.info_for_open_file,
                                                                             self.info_path_open_file,
                                                                             self.info_extention_open_file)
        # выбор только пути файла из data_of_open_file_name
        file_name = data_of_open_file_name[0]

        # выбор где и что менять исходя из выбора пользователя
        # нажата кнопка выбора Полного файла
        if self.sender().objectName() == self.toolButton_select_full_file.objectName():
            if file_name == '':
                self.label_path_full_file.setText(old_path_of_selected_full_file)
                self.label_path_full_file.adjustSize()
            else:
                old_path_of_selected_full_file = self.label_path_full_file.text()
                self.label_path_full_file.setText(file_name)
                self.label_path_full_file.adjustSize()

        # нажата кнопка выбора Неполного файла
        if self.sender().objectName() == self.toolButton_select_half_file.objectName():
            if file_name == '':
                self.label_path_half_file.setText(old_path_of_selected_half_file)
                self.label_path_half_file.adjustSize()
            else:
                old_path_of_selected_half_file = self.label_path_half_file.text()
                self.label_path_half_file.setText(file_name)
                self.label_path_half_file.adjustSize()

        # активация и деактивация объектов на форме зависящее от выбраны ли все файлы и они разные
        if self.label_path_full_file.text() != self.label_path_half_file.text():
            # если выбранные файлы разные
            if self.text_empty_path_file not in (self.label_path_full_file.text(), self.label_path_half_file.text()):
                self.pushButton_do_fill_data.setEnabled(True)
                self.listWidget_specialization.setEnabled(True)
        else:
            # если выбранные файлы одинаковые
            self.pushButton_do_fill_data.setEnabled(False)
            self.listWidget_specialization.setEnabled(False)

        # очистка списка специализаций при любой смене файла
        self.listWidget_specialization.clear()

        # заполнение listWidget_specialization специальностями из Полного файла
        if self.listWidget_specialization.isEnabled():
            # открыть файлы Полный и Неполный, и выбрать листы
            wb_full = openpyxl.load_workbook(self.label_path_full_file.text())
            wb_full_s = wb_full.active
            wb_half = openpyxl.load_workbook(self.label_path_half_file.text())
            wb_half_s = wb_half.active

            # посчитать количество строк и вывести на форме
            # TODO
            # сделать правильное добавление - замена только числа количества строк, а не добавления строки с номером
            self.label_full_file.setText(self.label_full_file.text() + f' (строк в файле {str(wb_full_s.max_row -1)})')
            self.label_full_file.adjustSize()
            self.label_half_file.setText(self.label_half_file.text() + f' (строк в файле {str(wb_half_s.max_row -1)})')
            self.label_half_file.adjustSize()

            # сформированные диапазоны обработки
            range_full_file = self.range_all_files + wb_full_s.cell(wb_full_s.max_row, wb_full_s.max_column).coordinate
            wb_full_range = wb_full_s[range_full_file]

            # множество для хранения специальностей
            specialization_set = set()

            # цикл прохода по Полному файлу и взятие непустых специальностей
            for row_in_range_full in wb_full_range:
                if row_in_range_full[-1].value:
                    specialization_set.add(row_in_range_full[-1].value)

            # добавление в список отсортированных специальностей
            self.listWidget_specialization.addItems(sorted(specialization_set, reverse=False))
            wb_full.close()
            wb_half.close()

    # событие - нажатие на кнопку заполнения файла
    def do_fill_data(self):
        # TODO
        # сделать проверку lineEdit_max_string на число

        # выбор выбранных строк в списке специальностей
        specialization_selected = [item.text() for item in self.listWidget_specialization.selectedItems()]

        # проверка на количество выбранных строк в listWidget_specialization
        if len(specialization_selected) == 0:
            # информационное окно о сохранении файлов
            self.window_info = PyQt5.QtWidgets.QMessageBox()
            self.window_info.setWindowTitle('Выберите специальности')
            self.window_info.setText(f'В списке специальностей ничего не выбрано,\n'
                                     f'выберите хотя бы одну строку')
            self.window_info.exec_()
        else:
            # считаю время заполнения
            time_start = time.time()

            # открыть файл Полный и Неполный, и выбрать листы
            wb_full = openpyxl.load_workbook(self.label_path_full_file.text())
            wb_full_s = wb_full.active
            wb_half = openpyxl.load_workbook(self.label_path_half_file.text())
            wb_half_s = wb_half.active

            # сформированные диапазоны обработки
            range_full_file = self.range_all_files + wb_full_s.cell(wb_full_s.max_row, wb_full_s.max_column).coordinate
            range_half_file = self.range_all_files + wb_half_s.cell(wb_half_s.max_row, wb_half_s.max_column).coordinate
            wb_full_range = wb_full_s[range_full_file]
            wb_half_range = wb_half_s[range_half_file]

            # список одной строки прохода, список выбранных строк по специальностям, списки всех строк Неполного файла
            list_one_string = []  # временная переменная
            list_half_file = []  # весь Неполный файл
            list_filtered_string = []  # фильтрованные строки из Полного которые устраивают выбранным специальностям
            list_for_add = []  # список выбранных из фильтрованных для добавления в Неполный файл
            tuple_half_file = ()  # кортеж для хранения ФИО из Неполного файла

            # счётчик удачных добавлений в Неполный из выбранных строк
            count_add_succes = 0

            # заполнение list_half_file Неполного файла
            for row_in_range_half in wb_half_range:
                # чищу список для временной строки
                list_one_string = []

                # прохожу строку
                for cell_in_row_half in row_in_range_half:
                    list_one_string.append(cell_in_row_half.value)

                # все записи из Неполного файла
                list_half_file.append(list_one_string)

            # кортеж из Неполного файла с первыми тремя ячейками ФИО
            # для проверки вхождения случайно выбранного из list_filtered_string
            list_one_string = []
            for str_half in list_half_file:
                list_one_string.append(get_fio_low_case(str_half))
            tuple_half_file = tuple(list_one_string)

            # заполнение list_filtered_string фильтрованных из specialization_selected из Полного файла
            for row_in_range_full in wb_full_range:
                # чищу список для временной строки
                list_one_string = []

                # прохожу строку
                for cell_in_row_full in row_in_range_full:
                    list_one_string.append(cell_in_row_full.value)

                # если последнее значение в списке специальностей, то добавляю его в список выбранных из Полного файла
                if list_one_string[-1] in specialization_selected:
                    # проверка на вхожесть фильтрованного в Неполный файл
                    if get_fio_low_case(list_one_string) not in tuple_half_file:
                        list_filtered_string.append(list_one_string)

            # количество строк "сколько хочу строк" (перевод значения в поле шага 3)
            count_string_want = int(self.lineEdit_max_string.text())

            # количество строк в Неполном файле (-1 потому что верхняя строка это шапка)
            count_string_half = wb_half_s.max_row - 1

            # сколько нужно добавить строк в Неполный файл, должно быть больше нуля
            count_string_add = count_string_want - count_string_half

            # количество строк в отфильтрованном списке
            count_filter_string = len(list_filtered_string)

            # количество строк которых будет реально добавлены в Неполный файл
            count_real_data_add = count_filter_string - count_string_add

            # добавление строк в Неполный файл
            # если количество строк в Неполном меньше, чем хочется, то добавить разницу строк
            if count_string_add <= 0:
                # информационное окно
                self.window_info = PyQt5.QtWidgets.QMessageBox()
                self.window_info.setWindowTitle('Строки')
                self.window_info.setText(f'Количество строк в Неполном файле больше или одинаково,\n'
                                         f'чем в ПУНКТЕ 3, их разница равна {count_string_add}\n'
                                         f' \n'
                                         f'хочется чтобы было {count_string_want}\n'
                                         f'сейчас в файле {count_string_half}\n'
                                         f'надо добавить {count_string_add}\n'
                                         f'могу выбрать из {count_filter_string}'
                                         )
                self.window_info.exec_()
            else:
                if count_string_add > count_filter_string:
                    # если добавляемых больше, чем отфильтрованных, то добавлять всё из list_filtered_string
                    # информационное окно
                    self.window_info = PyQt5.QtWidgets.QMessageBox()
                    self.window_info.setWindowTitle('Строки')
                    self.window_info.setText(f'Количество строк в Полном файле по этим специальностям\n'
                                             f'меньше, чем в ПУНКТЕ 3, их разница равна {count_real_data_add}\n'
                                             f'выберите ещё специальностей из списка\n'
                                             f' \n'
                                             f'хочется чтобы было {count_string_want}\n'
                                             f'сейчас в файле {count_string_half}\n'
                                             f'надо добавить {count_string_add}\n'
                                             f'могу выбрать из {count_filter_string}')
                    self.window_info.exec_()
                else:
                    # выбираю count_string_add штук в список случайных строк из фильтрованного
                    list_for_add = random.sample(list_filtered_string, count_string_add)

                    # последняя строка в Неполном +2 потому, что один за прошлый вычет, а один на следующую строчку
                    string_half_begin = (count_string_half + 1) + 1
                    string_half_end = (count_string_half + 1) + len(list_for_add)

                    # TODO
                    # добавление данных в эксель
                    for string_list_for_add in list_for_add:
                        wb_half_s.append(string_list_for_add)

                    # сохраняю файл и закрываю оба
                    filename_half = os.path.split(self.label_path_half_file.text())[1]
                    wb_half.save(filename_half)
                    wb_full.close()
                    wb_half.close()

                    # считаю время заполнения
                    time_finish = time.time()

                    # информационное окно о сохранении файлов
                    self.window_info = PyQt5.QtWidgets.QMessageBox()
                    self.window_info.setWindowTitle('Файл')
                    self.window_info.setText(f'Файлы сохранены и закрыты.\n'
                                             f'Заполнение сделано за {round(time_finish - time_start, 1)} секунд')
                    self.window_info.exec_()

    # событие - нажатие на кнопку Выход
    @staticmethod
    def click_on_btn_exit():
        exit()


# создание основного окна
def main_app():
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    app_window_main = Window()
    app_window_main.show()
    sys.exit(app.exec_())


# запуск основного окна
if __name__ == '__main__':
    main_app()

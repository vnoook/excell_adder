import time
import openpyxl

# считаю время скрипта
time_start = time.time()
print('начинается' + '.'*20)

# файлы для работы
xl_my_with_id = 'res/25_26_11_2021_my_with_id.xlsx'
xl_real_data = 'res/25_11_2021_real_data.xlsx'
xl_ext_data = 'res/25-26.11.2021-ext_data.xlsx'
xl_all_users = 'res/allusers.xlsx'

# переменные для работы
min_row_xl_my_with_id = 2
max_row_xl_my_with_id = 1249
min_col_xl_my_with_id = 1
max_col_xl_my_with_id = 23

min_row_xl_real_data = 2
max_row_xl_real_data = 11561
min_col_xl_real_data = 1
max_col_xl_real_data = 12

min_row_xl_ext_data = 2
max_row_xl_ext_data = 1246
min_col_xl_ext_data = 1
max_col_xl_ext_data = 4  # потому что нужно только 25 число

min_row_xl_all_users = 2
max_row_xl_all_users = 11561
min_col_xl_all_users = 1
max_col_xl_all_users = 12

# открываю книги
wb_my_with_id = openpyxl.load_workbook(xl_my_with_id)
wb_my_with_id_s = wb_my_with_id.active
wb_real_data = openpyxl.load_workbook(xl_real_data)
wb_real_data_s = wb_real_data.active
wb_ext_data = openpyxl.load_workbook(xl_ext_data)
wb_ext_data_s = wb_ext_data.active
wb_all_users = openpyxl.load_workbook(xl_all_users)
wb_all_users_s = wb_all_users.active

# алгоритмы добавления данных
for i in range(min_row_xl_my_with_id, max_row_xl_my_with_id+1):
    user_id = wb_my_with_id_s.cell(i, 1).value
    user_grade = 'через поиск совпадений в xl_all_users'
    user_title = 'через поиск совпадений в xl_all_users'
    user_date_reg = 'нет такой информации, надо выдумать алгоритм подсчёта от даты мероприятия в xl_real_data'
    user_platform = 'xl_real_data'
    user_country = 'xl_real_data'
    user_city = 'xl_real_data'
    user_ip = 'xl_real_data'
    user_time_in = 'xl_ext_data'
    user_time_out = 'xl_ext_data'
    user_duration = 'xl_ext_data'
    user_percent = 'xl_ext_data'

    print(i, ' == ', user_id)

# for dict_key in xl_pm_sheets:
#     if wb_pm.index(wb_pm_s) in (1, 2, 3, 4, 5, 6):
#         for i_row in range(9, max_row_first_page+1):
#             for i_col in range(2, max_col_first_page+1):
#                 # B9:AH60 -> B9:AH60 || R9C2:R60C34 -> R9C2:R60C34
#                 wb_pm_s.cell(i_row, i_col).value = wb_file_data_s.cell(i_row, i_col).value


# сохраняю файл и закрываю его
wb_my_with_id.save(xl_my_with_id)
wb_my_with_id.close()

# закрываю файлы из которых беру данные
wb_real_data.close()
wb_ext_data.close()
wb_all_users.close()


# считаю время скрипта
time_finish = time.time()
print('\n' + '.'*30 + 'закончено за', round(time_finish-time_start, 3), 'секунд')

# закрываю программу
# input('\nНажмите ENTER')

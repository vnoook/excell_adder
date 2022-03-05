import time
import openpyxl

# считаю время скрипта
time_start = time.time()
print('начинается' + '.'*20)


def get_parts(fio_str):
    fio_parts = fio_str.split(' ', 3)  # (' '.join(fio_str.split())).split()
    return fio_parts


# файл для работы
xl_file = 'res/spisok_main1.xlsx'

# открываю книгу и листы
wb_file = openpyxl.load_workbook(xl_file)
# wb_file_s = wb_file.active
wb_file_s1 = wb_file['Лист1']
wb_file_s2 = wb_file['Лист2']

# переменные для работы
max_row_s1 = wb_file_s1.max_row
end_male = ['вич']  # ['', '', '', '', '']
end_female = ['вна']

# алгоритм обработки данных
for i in range(2, max_row_s1 + 1):
    user_fio_s1 = wb_file_s1.cell(i, 1).value
    user_email_s1 = wb_file_s1.cell(i, 6).value

    print(f'{i} == {get_parts(user_fio_s1)} == {user_email_s1}')

    if get_parts(user_fio_s1)[0]:
        wb_file_s2.cell(i, 1).value = get_parts(user_fio_s1)[0]

    if get_parts(user_fio_s1)[1]:
        wb_file_s2.cell(i, 2).value = get_parts(user_fio_s1)[1]

    if len(get_parts(user_fio_s1)) > 2:
        wb_file_s2.cell(i, 3).value = get_parts(user_fio_s1)[2]

        part_word = get_parts(user_fio_s1)[2][-3:].lower()
        if part_word in end_male:
            wb_file_s2.cell(i, 4).value = 'муж'
        elif part_word in end_female:
            wb_file_s2.cell(i, 4).value = 'жен'
        else:
            wb_file_s2.cell(i, 4).value = '---'

    wb_file_s2.cell(i, 5).value = user_email_s1

# сохраняю файл и закрываю его
wb_file.save(xl_file)
wb_file.close()

# считаю время скрипта
time_finish = time.time()
print('\n' + '.'*30 + 'закончено за', round(time_finish-time_start, 3), 'секунд')

# закрываю программу
# input('\nНажмите ENTER')
